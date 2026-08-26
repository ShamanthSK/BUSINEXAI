from fastapi import APIRouter, Response
from pydantic import BaseModel
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

from app.api.dataset_store import get_dataset
from app.engine.what_if_engine import run_what_if_simulation

router = APIRouter()

class WhatIfRequest(BaseModel):
    marketing_change_pct: float = 0.0
    price_change_pct: float = 0.0
    conversion_change_pct: float = 0.0

@router.post("/datasets/{dataset_id}/what-if")
def simulate_what_if(dataset_id: str, req: WhatIfRequest):
    df = get_dataset(dataset_id)
    return run_what_if_simulation(
        df,
        marketing_change_pct=req.marketing_change_pct,
        price_change_pct=req.price_change_pct,
        conversion_change_pct=req.conversion_change_pct
    )

@router.post("/datasets/{dataset_id}/what-if/export/excel")
def export_what_if_excel(dataset_id: str, req: WhatIfRequest):
    df = get_dataset(dataset_id)
    sim = run_what_if_simulation(
        df,
        marketing_change_pct=req.marketing_change_pct,
        price_change_pct=req.price_change_pct,
        conversion_change_pct=req.conversion_change_pct
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "What-If Scenario Simulation"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    section_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)

    # Title Banner
    ws.merge_cells("A1:F1")
    t_cell = ws["A1"]
    t_cell.value = "STRATOS AI — STRATEGIC WHAT-IF SCENARIO REPORT"
    t_cell.font = title_font
    t_cell.fill = header_fill
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # Section 1: Simulated Assumptions
    ws.merge_cells("A3:F3")
    s1 = ws["A3"]
    s1.value = "1. SIMULATED ASSUMPTIONS"
    s1.font = section_font
    s1.fill = section_fill
    s1.alignment = Alignment(indent=1, vertical="center")

    assumptions = [
        ("Marketing Spend Shift (%)", f"{req.marketing_change_pct}%"),
        ("Price Adjustment (%)", f"{req.price_change_pct}%"),
        ("Conversion Velocity Change (%)", f"{req.conversion_change_pct}%")
    ]
    for r_idx, (k, v) in enumerate(assumptions, start=4):
        c1 = ws.cell(row=r_idx, column=1, value=k)
        c2 = ws.cell(row=r_idx, column=2, value=v)
        c1.font = bold_font
        c2.font = regular_font
        c1.border = cell_border
        c2.border = cell_border

    # Section 2: Outcome Summary
    ws.merge_cells("A8:F8")
    s2 = ws["A8"]
    s2.value = "2. SCENARIO OUTCOME SUMMARY"
    s2.font = section_font
    s2.fill = section_fill
    s2.alignment = Alignment(indent=1, vertical="center")

    summary_data = [
        ("Baseline Revenue", sim["baseline"]["revenue_formatted"]),
        ("Projected Revenue", f"{sim['projected']['revenue_formatted']} ({'+' if sim['projected']['revenue_change_pct'] >= 0 else ''}{sim['projected']['revenue_change_pct']}%)"),
        ("Baseline Profit", sim["baseline"]["profit_formatted"]),
        ("Projected Profit", f"{sim['projected']['profit_formatted']} ({'+' if sim['projected']['profit_change_pct'] >= 0 else ''}{sim['projected']['profit_change_pct']}%)"),
        ("Projected Net Profit Margin", f"{sim['projected']['profit_margin']}%"),
        ("Projected Active Accounts", str(sim["projected"]["customers"])),
        ("Expected Marketing ROI", f"{sim['projected']['expected_roi']}%")
    ]
    for r_idx, (k, v) in enumerate(summary_data, start=9):
        c1 = ws.cell(row=r_idx, column=1, value=k)
        c2 = ws.cell(row=r_idx, column=2, value=v)
        c1.font = bold_font
        c2.font = regular_font
        c1.border = cell_border
        c2.border = cell_border

    # Section 3: Monthly Trajectory Table
    ws.merge_cells("A17:F17")
    s3 = ws["A17"]
    s3.value = "3. MONTHLY TRAJECTORY BREAKDOWN"
    s3.font = section_font
    s3.fill = section_fill
    s3.alignment = Alignment(indent=1, vertical="center")

    headers = ["Month", "Baseline Revenue (₹)", "Projected Revenue (₹)", "Revenue Delta (₹)"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=18, column=c_idx, value=h)
        cell.font = bold_font
        cell.fill = card_fill
        cell.border = cell_border

    chart_data = sim.get("chart_data", [])
    for r_idx, row in enumerate(chart_data, start=19):
        base_rev = round(float(row.get("baseline_revenue", 0)), 2)
        proj_rev = round(float(row.get("projected_revenue", 0)), 2)
        delta = round(proj_rev - base_rev, 2)
        
        vals = [row.get("month", ""), base_rev, proj_rev, delta]
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = cell_border

    # Auto-adjust column dimensions
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 22

    # Embedded Excel Line Chart (Baseline vs Projected Revenue) positioned at H3
    if chart_data:
        chart = LineChart()
        chart.title = "Baseline vs Simulated Projected Revenue"
        chart.style = 13
        chart.y_axis.title = "Revenue (₹)"
        chart.x_axis.title = "Month"
        chart.width = 16
        chart.height = 10

        data_ref = Reference(ws, min_col=2, min_row=18, max_col=3, max_row=18 + len(chart_data))
        cats_ref = Reference(ws, min_col=1, min_row=19, max_row=18 + len(chart_data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "H3")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"STRATOS_WhatIf_Simulation_{dataset_id}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

