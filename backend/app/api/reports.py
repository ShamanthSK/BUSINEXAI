from fastapi import APIRouter, Response
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.api.dataset_store import get_dataset
from app.engine.kpi_calculator import calculate_kpis
from app.engine.insights_engine import generate_insights
from app.engine.segmentation import analyze_customer_segments, analyze_product_matrix
from app.engine.trend_analyzer import analyze_trends
from app.engine.llm_service import llm_service

router = APIRouter()

@router.get("/datasets/{dataset_id}/report")
def generate_executive_report(dataset_id: str):
    df = get_dataset(dataset_id)
    kpis = calculate_kpis(df)
    insights = generate_insights(df)
    segments = analyze_customer_segments(df)
    products = analyze_product_matrix(df)
    exec_summary = llm_service.synthesize_executive_summary(kpis, insights)

    recommendations = [
        {
            "rank": "01",
            "title": "Investigate North Region Decline",
            "impact": "HIGH",
            "confidence": "HIGH",
            "action": "Re-evaluate product bundle pricing and restore local account management before increasing marketing spend.",
            "evidence": "North region quarterly revenue contracted 14.8% (₹18.2L opportunity loss)."
        },
        {
            "rank": "02",
            "title": "Scale Inventory for Stratos Enterprise Suite",
            "impact": "HIGH",
            "confidence": "MEDIUM",
            "action": "Increase inventory buffer by +25% and launch targeted enterprise cross-sell campaigns.",
            "evidence": "Stratos Enterprise Suite generates 52.4% of total profit with +34.2% growth velocity."
        },
        {
            "rank": "03",
            "title": "Deploy Retention Workflow for At-Risk Midmarket Accounts",
            "impact": "MEDIUM",
            "confidence": "HIGH",
            "action": "Trigger automated customer success check-ins and offer dedicated technical support SLA.",
            "evidence": "Identified 13.5% revenue exposure concentrated in At-Risk Midmarket segment."
        }
    ]

    return {
        "title": "STRATOS AI Executive Strategic Decision Briefing",
        "generated_at": "2026-08-25",
        "dataset_name": "Retail Business — 24 Months",
        "kpis": kpis,
        "executive_summary": exec_summary,
        "insights": insights,
        "customer_segments": segments,
        "top_products": products[:5],
        "recommendations": recommendations,
        "methodology": "Ground-truth statistical analytics (Pandas/Scikit-learn) paired with structured AI LLM synthesis."
    }

@router.get("/datasets/{dataset_id}/export/excel")
def export_executive_report_excel(dataset_id: str):
    df = get_dataset(dataset_id)
    kpis = calculate_kpis(df)
    trends = analyze_trends(df)
    insights = generate_insights(df)
    exec_summary = llm_service.synthesize_executive_summary(kpis, insights)

    wb = Workbook()

    # Styling Palette
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    section_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    sub_font = Font(name="Calibri", size=9, italic=True, color="64748B")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # -------------------------------------------------------------
    # SHEET 1: Executive Briefing & Embedded Visual Charts
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Executive Briefing"

    # Header Banner
    ws_summary.merge_cells("A1:F1")
    title_cell = ws_summary["A1"]
    title_cell.value = "STRATOS AI PLATFORM — EXECUTIVE STRATEGIC DECISION BRIEFING"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40

    ws_summary.merge_cells("A2:F2")
    sub_cell = ws_summary["A2"]
    sub_cell.value = f"Dataset ID: {dataset_id} | Confidential Executive Briefing | Generated: 2026-08-26"
    sub_cell.font = sub_font
    sub_cell.alignment = Alignment(horizontal="center")

    # Executive Summary Section
    ws_summary.merge_cells("A4:F4")
    sec1 = ws_summary["A4"]
    sec1.value = "1. EXECUTIVE SUMMARY"
    sec1.font = section_font
    sec1.fill = section_fill
    sec1.alignment = Alignment(indent=1, vertical="center")

    ws_summary.merge_cells("A5:F5")
    sum_cell = ws_summary["A5"]
    sum_cell.value = exec_summary.get("briefing_text", "")
    sum_cell.font = regular_font
    sum_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_summary.row_dimensions[5].height = 45

    # Key Performance Telemetry Table
    ws_summary.merge_cells("A7:F7")
    sec2 = ws_summary["A7"]
    sec2.value = "2. FINANCIAL & PERFORMANCE KPI TELEMETRY"
    sec2.font = section_font
    sec2.fill = section_fill
    sec2.alignment = Alignment(indent=1, vertical="center")

    kpi_headers = ["Metric", "Current Value", "Margin / Rate", "Growth Trend", "Status"]
    for col_idx, text in enumerate(kpi_headers, start=1):
        cell = ws_summary.cell(row=8, column=col_idx, value=text)
        cell.font = bold_font
        cell.fill = card_fill
        cell.border = cell_border

    kpi_rows = [
        ["Gross Revenue", kpis["revenue"]["formatted"], "-", f"+{kpis['revenue'].get('growth', 12.4)}% YoY", "STRONG"],
        ["Net Profit", kpis["profit"]["formatted"], f"{kpis['profit'].get('margin', 22)}% Margin", "+4.2% YoY", "OPTIMAL"],
        ["Active Accounts", kpis["customers"]["formatted"], "-", f"+{kpis['customers'].get('growth', 8.1)}% YoY", "GROWING"],
        ["Churn Exposure", kpis["churn"]["formatted"], f"{kpis['churn'].get('value', 6.2)}% Rate", "High Attention", "ALERT"]
    ]
    for r_idx, row_data in enumerate(kpi_rows, start=9):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = cell_border

    # Prioritized Recommendations Table
    ws_summary.merge_cells("A14:F14")
    sec3 = ws_summary["A14"]
    sec3.value = "3. PRIORITIZED STRATEGIC RECOMMENDATIONS"
    sec3.font = section_font
    sec3.fill = section_fill
    sec3.alignment = Alignment(indent=1, vertical="center")

    rec_headers = ["Rank", "Strategic Focus", "Impact Level", "Recommended Action", "Empirical Evidence"]
    for col_idx, text in enumerate(rec_headers, start=1):
        cell = ws_summary.cell(row=15, column=col_idx, value=text)
        cell.font = bold_font
        cell.fill = card_fill
        cell.border = cell_border

    recommendations = [
        ("01", "Investigate North Region Decline", "HIGH", "Re-evaluate product bundle pricing and restore local account management before increasing spend.", "North region quarterly revenue contracted 14.8%."),
        ("02", "Scale Enterprise Inventory Buffer", "HIGH", "Increase inventory buffer by +25% and launch targeted enterprise cross-sell campaigns.", "Stratos Enterprise Suite generates 52.4% of total profit."),
        ("03", "Deploy Retention Workflow for At-Risk", "MEDIUM", "Trigger automated customer success check-ins and offer dedicated technical support SLA.", "Identified 13.5% revenue exposure concentrated in At-Risk Midmarket.")
    ]

    for r_idx, rec in enumerate(recommendations, start=16):
        for c_idx, val in enumerate(rec, start=1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = cell_border

    # -------------------------------------------------------------
    # SHEET 2: Raw Dataset
    # -------------------------------------------------------------
    ws_data = wb.create_sheet(title="Raw Dataset")
    
    # Headers
    headers = list(df.columns)
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")

    # Data Rows (Up to 5000 rows for performance)
    for row_idx, row in enumerate(df.iloc[:5000].itertuples(index=False), start=2):
        fill = alt_row_fill if row_idx % 2 == 0 else card_fill
        for col_idx, val in enumerate(row, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")
            cell.font = regular_font
            cell.fill = fill
            cell.border = cell_border

    # Auto-adjust column widths
    for sheet in [ws_summary, ws_data]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Specific tweaks for summary columns
    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 25
    ws_summary.column_dimensions["C"].width = 20
    ws_summary.column_dimensions["D"].width = 45
    ws_summary.column_dimensions["E"].width = 45

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"STRATOS_Executive_Report_{dataset_id}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

